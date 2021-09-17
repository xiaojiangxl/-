from openpyxl import  workbook,load_workbook    #操作excel
import requests,re,time
import dingtalkchatbot.chatbot as cb
import datetime

def vul_read():

    xlsx_name = '活动信息汇总.xlsx'
    wb = load_workbook(xlsx_name)  # 打开工作簿
    sheet = wb['Sheet1']  # 这是我们的需要的表,请确保该表的名字是默认
    stat = 1  # 设置开关
    singe_vul = []   #临时储存单条漏洞信息
    vuls = []     #储存所有信息
    for row in sheet.iter_rows(min_row=2,max_row=1000, max_col=10):  # 遍历,读取区域
        for cel in row:
            if cel.value == None:  # 没数据了，就退出
                break
            singe_vul.append(cel.value)    #将每一条的每个单元格的数据存进去
        if singe_vul != [] :         #去掉空的列表
            vuls.append(singe_vul)
        singe_vul = []             #只能用赋值的形式清空列表，如果有clear，就会导致逻辑错误，最终的结果都是空
        if stat == 0:
            break
    return vuls

def dingding(text, msg):
    # 将此处换为钉钉机器人的api
    webhook = 'https://oapi.dingtalk.com/robot/send?access_token='
    ding = cb.DingtalkChatbot(webhook)
    ding.send_text(msg='{}\r\n{}'.format(text, msg), is_at_all=False) 

# server酱  http://sc.ftqq.com/?c=code
def server(text, msg):
    # 将 xxxx 换成自己的server SCKEY
    ii='https://sc.ftqq.com/SCU1.send'+'?text={}&desp={}'
    uri = ii.format(text, msg)
    send = requests.get(uri)

def shijian(Ti):
    ti=time.strftime('%Y.%m.%d')
    ti=datetime.datetime.strptime(ti,"%Y.%m.%d")
    Ti=datetime.datetime.strptime(Ti,"%Y.%m.%d")
    if(ti<Ti):#还没到日期
        return 1
    elif(ti==Ti):#当天
        return 0
    elif(ti>Ti):#已经过了
        return -1

def log(text):
    f=open('./log.txt',mode='a',encoding='utf-8')
    f.write(str(text)+'\n')
    print(str(text))

def tongzhi(text,msg):
    #server(text,msg)
    dingding(text,msg)

def panduan():
    message = vul_read()
    for result in message:
        kaishi=shijian(result[1])
        jieshu=shijian(result[2])
        if(kaishi<1 and jieshu>-1):
            text=str(result[0])+'活动推送'
            msg='活动截止日期：'+str(result[2])+'\n活动简介：'+str(result[3])+'\n活动链接: '+str(result[4])
            tongzhi(text,msg)
            logtext=str(time.strftime('%Y.%m.%d,%H:%M:%S'))+' [发送成功]'+str(result[0])+'活动推送'
            log(logtext)
            time.sleep(30)
if __name__ == '__main__':
    print('正在推送中....')
    while True:
       
        try:
            tim=time.strftime('%H:%M')
            if(tim=='09:00' or tim=='20:30'):
                panduan()
                time.sleep(61)
        except:
            pass
